from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date

def generate_equipments_report(equipment_data):
    wb = Workbook()
    ws = wb.active

    ws.title = "controle_equipamentos"

    headers = [
        "Nome", "Tipo", "Marca", "Modelo", "Nº de Série",
        "Setor", "Localização", "Status", "Data de Aquisição",
        "Intervalo de Manutenção (meses)", "Grupo de Manutenção"
    ]

    last_column_letter = get_column_letter(len(headers))

    # Titulo
    ws.merge_cells(f"A1:{last_column_letter}1")

    title_cell = ws["A1"]
    title_cell.value = "Relatório de Equipamentos"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Dados
    for row_index, equipment in enumerate(equipment_data, start=3):
        acquisition_date = equipment.get("acquisition_date")

        if isinstance(acquisition_date, (datetime, date)):
            acquisition_date = acquisition_date.strftime("%d/%m/%Y")
        elif acquisition_date:
            acquisition_date = str(acquisition_date)

        row_values = [
            equipment.get("name"),
            equipment.get("type"),
            equipment.get("brand"),
            equipment.get("model"),
            equipment.get("serial_number"),
            equipment.get("sector"),
            equipment.get("location"),
            equipment.get("status"),
            acquisition_date,
            equipment.get("maintenance_interval_months"),
            equipment.get("maintenance_group"),
        ]

        for col_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajusta a largura das colunas de acordo com o maior conteúdo de cada uma
    for col_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_index)
        max_length = len(header)

        for row in ws.iter_rows(min_row=3, min_col=col_index, max_col=col_index):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 4

    ws.freeze_panes = "A3"

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream
